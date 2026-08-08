import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Load the original workbook
workbook = xl.load_workbook('data/transactions.xlsx')
worksheet = workbook['Sheet1']

# Update column D based on column C
for row in range(2, worksheet.max_row + 1):
    price_cell = worksheet.cell(row, 3)
    corrected_price = price_cell.value * 0.9

    corrected_price_cell = worksheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Create chart using column D
ref = Reference(
    worksheet,
    min_row=2,
    max_row=worksheet.max_row,
    min_col=4,
    max_col=4
)

bar_chart = BarChart()
bar_chart.add_data(ref)

worksheet.add_chart(bar_chart, "E1")

# Save as a NEW file
workbook.save('data/transactions2.xlsx')